#!/usr/bin/env python3
"""Kit de revue portable des cartes finales (à envoyer sur un autre PC).

1. revue-cartes-gacha.html — autonome : les 2304 miniatures EMBARQUÉES en
   base64, groupées par collection, avec nom/rareté/PV/source d'image, lien
   vers la page Wikipedia et vers l'image pleine taille en ligne (GitHub Pages).
2. revue-cartes-gacha.xlsx — feuille de corrections : une ligne par carte,
   colonnes à remplir (Nom corrigé / Lien nouvelle image / Remarques).
"""
import sys, json, base64, html as h
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
SORTIE_HTML = Path(r'C:\Users\flxjr\Downloads\revue-cartes-gacha.html')
SORTIE_XLSX = Path(r'C:\Users\flxjr\Downloads\revue-cartes-gacha.xlsx')
BASE_URL = 'https://blinytz.github.io/gacha-wikipedia/'

COULEUR_RARETE = {'commune': '#8a93a6', 'rare': '#4aa8ff', 'epique': '#b05cff',
                  'mythique': '#ff5cd0', 'legendaire': '#ffd166'}

index = json.loads((ROOT / 'data' / 'collections.json').read_text(encoding='utf-8'))
sources = json.loads((ROOT / 'build' / 'images_sources.json').read_text(encoding='utf-8'))
collections = []
for c in index['collections']:
    collections.append(json.loads((ROOT / c['fichier']).read_text(encoding='utf-8')))

# ---------------------------------------------------------------- HTML
parts = ["""<!doctype html><html lang="fr"><meta charset="utf-8">
<title>Revue des cartes — Gacha Wikipedia</title>
<style>
 body{background:#14141e;color:#eceaf6;font-family:system-ui,sans-serif;margin:16px}
 h1{font-size:20px} h2{font-size:16px;margin:26px 0 10px;position:sticky;top:0;
   background:#14141e;padding:8px 0;border-bottom:1px solid #333}
 nav{display:flex;flex-wrap:wrap;gap:6px;margin:12px 0}
 nav a{color:#46c8ff;text-decoration:none;font-size:12px;background:#22223300;
   border:1px solid #46c8ff44;border-radius:999px;padding:3px 10px}
 .g{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:10px}
 .c{background:#1e1e2c;border-radius:10px;padding:8px;border-top:3px solid var(--r)}
 .c img{width:100%;height:120px;object-fit:cover;border-radius:6px;background:#111}
 .n{font-size:12px;font-weight:700;margin:6px 0 2px;line-height:1.25}
 .p{font-size:10px;color:#9a97b0} .p b{color:var(--r)}
 .s{font-size:9px;color:#6e6b85}
 .l{font-size:10px} .l a{color:#46c8ff;text-decoration:none;margin-right:8px}
 .avert{background:#463a10;border:1px solid #ffd16660;border-radius:8px;
   padding:10px 14px;font-size:13px;margin:10px 0}
</style>
<h1>Revue des cartes — version finale (données réelles + images)</h1>
<div class="avert">Pour chaque carte : miniature réelle, nom de carte,
rareté/PV, source de l'image, lien [wiki] vers la page liée, lien [full]
vers l'image grand format en ligne (connexion requise pour [full] seulement).
Note tes corrections dans <b>revue-cartes-gacha.xlsx</b> (nom corrigé, lien
d'image de remplacement, remarques) et renvoie-le-moi.</div>"""]

parts.append('<nav>' + ''.join(
    f'<a href="#{d["slug"]}">{h.escape(d["collection"])} ({len(d["cartes"])})</a>'
    for d in collections) + '</nav>')

nb_img = 0
for d in collections:
    parts.append(f'<h2 id="{d["slug"]}">{h.escape(d["collection"])} '
                 f'({len(d["cartes"])} cartes)</h2><div class="g">')
    for carte in d['cartes']:
        fslug = carte['id'].split('_', 1)[1]
        chemin = ROOT / 'images' / 'thumbs' / d['slug'] / f'{fslug}.webp'
        if chemin.exists():
            b64 = base64.b64encode(chemin.read_bytes()).decode()
            img = f'<img src="data:image/webp;base64,{b64}" loading="lazy">'
            nb_img += 1
        else:
            img = '<img alt="MANQUANTE">'
        src = sources.get(carte['id'], {}).get('source', '?')
        couleur = COULEUR_RARETE[carte['rarete']]
        page = ('' if carte['nom'] == carte['titrePage']
                else f' <span class="s">(page : {h.escape(carte["titrePage"])})</span>')
        parts.append(f"""<div class="c" style="--r:{couleur}">{img}
 <div class="n">{h.escape(carte['nom'])}{page}</div>
 <div class="p"><b>{carte['rarete']}</b> · {carte['pv']} PV · {carte['pageviews']:,} vues</div>
 <div class="s">image : {h.escape(src)}</div>
 <div class="l"><a href="{h.escape(carte['lienWikipedia'])}" target="_blank">wiki ↗</a>
 <a href="{BASE_URL}{h.escape(carte['imageUrl'])}" target="_blank">full ↗</a></div></div>""")
    parts.append('</div>')
parts.append('</html>')
SORTIE_HTML.write_text('\n'.join(parts), encoding='utf-8')

# ---------------------------------------------------------------- XLSX
wb = openpyxl.Workbook()
ws0 = wb.active
ws0.title = 'Consignes'
consignes = [
    ('Revue finale des cartes — Gacha Wikipedia', 14, True),
    ('', 10, False),
    ('Ouvre revue-cartes-gacha.html (même dossier) pour voir toutes les images.', 10, False),
    ('Ici, remplis uniquement les colonnes jaunes pour ce qui doit changer :', 10, False),
    ('  • «Nom corrigé» : le titre de carte que tu veux voir affiché.', 10, False),
    ('  • «Lien nouvelle image» : URL d’une meilleure image (n’importe quelle source).', 10, False),
    ('  • «Remarques» : tout le reste (mauvaise page liée, carte à supprimer, etc.).', 10, False),
    ('Les lignes sans rien de rempli = validées telles quelles.', 10, False),
    ('Renvoie ce fichier — j’appliquerai les corrections une par une.', 10, False),
]
for i, (texte, taille, gras) in enumerate(consignes, 1):
    cel = ws0.cell(row=i, column=1, value=texte)
    cel.font = Font(name='Arial', size=taille, bold=gras)
ws0.column_dimensions['A'].width = 95

ws = wb.create_sheet('Cartes')
entetes = ['Collection', 'N°', 'Nom de carte', 'Page Wikipedia liée', 'Rareté',
           'PV', 'Vues 12 mois', 'Source image', 'Voir image', 'Page wiki',
           'Nom corrigé', 'Lien nouvelle image', 'Remarques']
larg = [24, 5, 26, 26, 11, 6, 12, 15, 10, 10, 24, 30, 34]
GRIS = PatternFill('solid', start_color='2B2B40')
JAUNE = PatternFill('solid', start_color='FFF2CC')
BORD = Border(bottom=Side(style='thin', color='DDDDDD'))
FILL_RARETE = {'commune': 'E3E6EB', 'rare': 'CFE6FF', 'epique': 'E8D5FF',
               'mythique': 'FFD5F2', 'legendaire': 'FFEFC2'}
for j, (titre, l) in enumerate(zip(entetes, larg), 1):
    cel = ws.cell(row=1, column=j, value=titre)
    cel.font = Font(name='Arial', bold=True, color='FFFFFF')
    cel.fill = PatternFill('solid', start_color='7F6000') if j >= 11 else GRIS
    cel.alignment = Alignment(vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = l

ligne = 2
for d in collections:
    for carte in d['cartes']:
        src = sources.get(carte['id'], {}).get('source', '?')
        vals = [d['collection'], carte['numero'], carte['nom'], carte['titrePage'],
                carte['rarete'], carte['pv'], carte['pageviews'], src,
                None, None, '', '', '']
        for j, v in enumerate(vals, 1):
            cel = ws.cell(row=ligne, column=j, value=v)
            cel.font = Font(name='Arial', size=9)
            cel.border = BORD
            cel.alignment = Alignment(vertical='top')
        rar = ws.cell(row=ligne, column=5)
        rar.fill = PatternFill('solid', start_color=FILL_RARETE[carte['rarete']])
        img = ws.cell(row=ligne, column=9, value='image ↗')
        img.hyperlink = BASE_URL + carte['imageUrl']
        wiki = ws.cell(row=ligne, column=10, value='wiki ↗')
        wiki.hyperlink = carte['lienWikipedia']
        for c in (img, wiki):
            c.font = Font(name='Arial', size=9, color='1155CC', underline='single')
        for j in (11, 12, 13):
            ws.cell(row=ligne, column=j).fill = JAUNE
        ligne += 1

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:M{ligne - 1}'
wb.save(SORTIE_XLSX)

taille_html = SORTIE_HTML.stat().st_size / 1e6
taille_xlsx = SORTIE_XLSX.stat().st_size / 1e6
print(f'{SORTIE_HTML.name} : {nb_img} miniatures embarquées, {taille_html:.1f} Mo')
print(f'{SORTIE_XLSX.name} : {ligne - 2} lignes, {taille_xlsx:.1f} Mo')
