#!/usr/bin/env python3
"""Fichier Excel portable de vérification des 2328 cartes (étape 1).

Une ligne par carte : résolution + statut coloré + colonne « Titre corrigé »
à remplir, puis à renvoyer pour intégration dans overrides.json.
Aucune formule : le fichier est purement des données, lisible partout.
"""
import json, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
SORTIE = Path(r'C:\Users\flxjr\Downloads\verification-cartes-gacha.xlsx')

FONT = 'Arial'
COULEURS = {
    'ok':       ('D9EAD3', '2E7D32'),   # fond vert clair / texte vert
    'verifier': ('FCE5CD', 'B45F06'),   # orange
    'echec':    ('F4CCCC', 'CC0000'),   # rouge
}
JAUNE = PatternFill('solid', start_color='FFF2CC')
GRIS_ENTETE = PatternFill('solid', start_color='2B2B40')
BORD = Border(bottom=Side(style='thin', color='DDDDDD'))

rs = json.loads((ROOT / 'build' / 'resolution.json').read_text(encoding='utf-8'))

wb = Workbook()

# ---------- Feuille consignes ----------
ws0 = wb.active
ws0.title = 'Consignes'
consignes = [
    ('Vérification des cartes — Gacha Wikipedia (étape 1)', 14, True),
    ('', 10, False),
    ('Chaque ligne = une carte. La résolution automatique a associé chaque nom '
     'de ta liste à un article de fr.wikipedia.org.', 10, False),
    ('', 10, False),
    ('Statuts :', 11, True),
    ('  • ok (vert) : lookup direct ou correspondance exacte — erreur peu probable.', 10, False),
    ('  • verifier (orange) : résolu par recherche contextuelle — à contrôler en priorité.', 10, False),
    ('  • echec (rouge) : page introuvable — correction obligatoire.', 10, False),
    ('', 10, False),
    ('Comment corriger :', 11, True),
    ('  1. Filtre la colonne Statut sur « verifier » (et regarde les notes DOUBLON).', 10, False),
    ('  2. Clique le lien Wikipedia pour vérifier que la page correspond au sujet voulu.', 10, False),
    ("  3. Si c'est faux : écris le TITRE EXACT de la bonne page dans « Titre corrigé »", 10, False),
    ("     (copie le titre depuis l'article Wikipedia, accents compris).", 10, False),
    ('  4. Pour un DOUBLON (même page dans 2 collections) : garde une seule ligne telle', 10, False),
    ("     quelle et, sur l'autre, écris un titre de remplacement ou « RETIRER ».", 10, False),
    ('  5. Renvoie ce fichier — les corrections seront intégrées à overrides.json.', 10, False),
    ('', 10, False),
    ('Astuce : le fichier audit.html (fourni séparément) montre les mêmes lignes avec les '
     'images — pratique pour repérer les erreurs d’un coup d’œil.', 10, False),
]
for i, (texte, taille, gras) in enumerate(consignes, 1):
    c = ws0.cell(row=i, column=1, value=texte)
    c.font = Font(name=FONT, size=taille, bold=gras)
ws0.column_dimensions['A'].width = 110

# ---------- Feuille cartes ----------
ws = wb.create_sheet('Cartes')
entetes = ['Collection', 'N°', 'Nom Excel', 'Titre résolu', 'Statut', 'Note',
           'Extrait (début)', 'Lien Wikipedia', 'Titre corrigé (à remplir)']
larg = [26, 5, 28, 30, 9, 34, 46, 34, 30]
for j, (h, l) in enumerate(zip(entetes, larg), 1):
    c = ws.cell(row=1, column=j, value=h)
    c.font = Font(name=FONT, bold=True, color='FFFFFF')
    c.fill = GRIS_ENTETE
    c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(j)].width = l
ws.cell(row=1, column=9).fill = PatternFill('solid', start_color='7F6000')

for i, r in enumerate(rs, 2):
    fond, texte = COULEURS[r['statut']]
    vals = [r['collection'], r['numero'], r['nom'], r['titre'], r['statut'],
            r['note'] or r['methode'], (r['extrait'] or '')[:180], None, '']
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(name=FONT, size=9)
        c.border = BORD
        c.alignment = Alignment(vertical='top', wrap_text=(j in (6, 7)))
    sc = ws.cell(row=i, column=5)
    sc.fill = PatternFill('solid', start_color=fond)
    sc.font = Font(name=FONT, size=9, bold=True, color=texte)
    if r['statut'] != 'ok':      # ligne teintée pour repérage au scroll
        for j in (1, 3, 4):
            ws.cell(row=i, column=j).fill = PatternFill('solid', start_color=fond)
    if r['url']:
        lc = ws.cell(row=i, column=8, value=r['titre'])
        lc.hyperlink = r['url']
        lc.font = Font(name=FONT, size=9, color='1155CC', underline='single')
    ws.cell(row=i, column=9).fill = JAUNE

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:I{len(rs) + 1}'

wb.save(SORTIE)
ok = sum(1 for r in rs if r['statut'] == 'ok')
ver = sum(1 for r in rs if r['statut'] == 'verifier')
ech = sum(1 for r in rs if r['statut'] == 'echec')
print(f'{SORTIE} — {len(rs)} lignes ({ok} ok, {ver} à vérifier, {ech} échecs)')
